# -*- coding: utf-8 -*-
import xlrd
import openpyxl

# Below are dictionaries for data transform

HorizontalAlignment03Dict = {0: 'Default', 1: 'Left', 2: 'Center', 3: 'Right', 5: 'Justified', 4: 'Filled',
                             7: 'Distributed'}
HorizontalAlignment07Dict = {None: 'Default', 'general': 'Default', 'left': 'Left', 'center': 'Center',
                             'right': 'Right',
                             'justify': 'Justified', 'fill': 'Filled', 'distributed': 'Distributed'}
VerticalAlignment03Dict = {0: 'Top', 1: 'Middle', 2: 'Bottom', 3: 'Bottom', 4: 'Distributed'}
# VerticalAlignment03Dict = {0: 'Top', 1: 'Middle', 2: 'Bottom', 3: 'Justified', 4: 'Distributed'}
VerticalAlignment07Dict = {None: 'Bottom', 'top': 'Top', 'center': 'Middle', 'bottom': 'Bottom',
                           'distributed': 'Distributed'}
# VerticalAlignment07Dict = {'top': 'Top', 'center': 'Middle', 'bottom': 'Bottom', 'bottom': 'Justified','distributed': 'Distributed'}
Border03Dict = {0: 'No', 1: 'Solid', 2: 'Solid', 5: 'Solid', 13: 'Dotted', 4: 'Dotted', 8: 'Dashed', 3: 'Dashed',
                13: 'Dashed', 9: 'DashDot', 10: 'DashDot', 11: 'DashDotDot', 12: 'DashDotDot', 6: 'DoubleThin'}
Border07Dict = {None: 'No', 'thin': 'Solid', 'dotted': 'Dotted', 'dashed': 'Dashed', 'dashDot': 'DashDot',
                'dashDotDot': 'DashDotDot', 'double': 'DoubleThin'}


# read CSV data to produce my own data format, but no font, color or cell format supports for CSV file
def readCSV(path):
    content = []  # represents one csv file, to store data of one or more sheets
    with open(path, 'r') as f:  # open file to read
        sheetcontent = []  # represents one sheet, to store rows
        for line in f:
            line = line.replace(' ', '')
            if line == '\n':
                continue
            rowcontent = []  # represents one row, to store cells
            line = line.replace('\n', '')
            celllist = line.split(',')
            for cell in celllist:
                cellcontent = {'value': cell, 'size': 10, 'bold': 0, 'italic': 0,
                               'underline': 0,
                               'fontname': 'Arial', 'valign': 'Bottom',
                               'halign': 'Default',
                               'tborder': 'No', 'bborder': 'No',
                               'lborder': 'No', 'rborder': 'No'}  # store a cell's data and format
                rowcontent.append(cellcontent)
            #print(line)
            sheetcontent.append(rowcontent)
    content.append([sheetcontent, path, []])
    #print(content)
    return content
    pass


# read xls( only for suites before office 2003) files' data value, font and cell format( alignment and border)
# to my own data format
def read03xls(path):
    workbook = xlrd.open_workbook(path, formatting_info=True)  # set 'True' for formatting_info
    # to be able to access cell format
    sheets = workbook.sheet_names()  # read sheets' names to 'sheets'
    content = []  # represents one csv file, to store data of one or more sheets
    for sheetname in sheets:  # read every single sheet
        #print('sheet: ', sheetname)
        worksheet = workbook.sheet_by_name(sheetname)  # get sheet by sheetname
        sheetcontent = []  # represents one sheet, to store rows
        for i in range(0, worksheet.nrows):  # get each row
            # row = worksheet.row(i)
            rowcontent = []  # represents one row, to store cells
            for j in range(0, worksheet.ncols):  # get each cell
                cell = worksheet.cell(i, j)  # get cell object
                # print(cell.value, '\t', end='')
                fmtindex = cell.xf_index
                fmt = workbook.xf_list[fmtindex]  # get cell format object by format index
                fontindex = fmt.font_index
                font = workbook.font_list[fontindex]  # get cell font object
                size = (font.height / 20)
                bold = font.bold
                italic = font.italic
                underline = font.underlined
                fontname = font.name
                # color = font.colour_index  # font color
                # bgx = fmt.background.pattern_colour_index  # background color
                align = fmt.alignment  # get alignment object
                valign = align.vert_align  # vertical alignment
                halign = align.hor_align  # horizontal alignment
                border = fmt.border
                lborder = border.left_line_style  # left border
                rborder = border.right_line_style  # right border
                tborder = border.top_line_style  # top border
                bborder = border.bottom_line_style  # bottom border
                cvalue = cell.value  # cell value
                if cell.value == None:  # skip None cell( blank cell with nothing)
                    cvalue = ''
                else:
                    cvalue == str(cell.value)
                cellcontent = {'value': cvalue, 'size': size, 'bold': bold, 'italic': italic,
                               'underline': underline,
                               'fontname': 'Arial', 'valign': VerticalAlignment03Dict.get(valign),
                               'halign': HorizontalAlignment03Dict.get(halign),
                               'tborder': Border03Dict.get(tborder), 'bborder': Border03Dict.get(bborder),
                               'lborder': Border03Dict.get(lborder), 'rborder': Border03Dict.get(rborder)}
                # store a cell's data and format
                rowcontent.append(cellcontent)
            sheetcontent.append(rowcontent)
            #print()
        # print(worksheet.merged_cells)
        merge = []  # store sheet's merged cells
        for m in worksheet.merged_cells:
            pass
            merge.append(((m[0], m[2]), (m[1] - 1, m[3] - 1)))
        # print(merge)
        content.append([sheetcontent, sheetname, merge])
    test = fmt.border.top_line_style
    # showDetail(test)
    #print(content)
    return content


# read xlsx( only for suites after office 2007)files' data value, font and cell format( alignment and border)
# to my own data format
def read07xlsx(path):
    workbook = openpyxl.load_workbook(path)
    sheetnames = workbook.sheetnames
    content = []  # represents one csv file, to store data of one or more sheets
    for sheetname in sheetnames:  # read every single sheet
        sheet = workbook[sheetname]
        #print('sheet: ', sheetname)
        sheetcontent = []  # represents one sheet, to store rows
        ii = jj = 1
        for row in sheet.rows:  # get each row
            rowcontent = []  # represents one row, to store cells
            jj = 1
            for cell in row:  # get each cell
                # print(cell.value, "\t", end="")
                cellfont = cell.font  # get font object
                # print(ii, jj)
                size = cellfont.sz
                bold = cellfont.b
                italic = cellfont.i
                fontname = cellfont.name
                underline = cellfont.underline
                # color = cellfont.color.rgb  # font color
                # bgx = cell.fill.fgColor.rgb  # background color
                align = cell.alignment  # alignment object
                valign = align.vertical  # vertical alignment
                halign = align.horizontal  # horizontal alignment
                border = cell.border
                lborder = border.left.style  # left border
                rborder = border.right.style  # right border
                tborder = border.top.style  # top border
                bborder = border.bottom.style  # bottom border
                cvalue = cell.value  # cell value
                if cell.value == None:  # skip None cell( blank cell with nothing)
                    cvalue = ''
                else:
                    cvalue == str(cell.value)
                cellcontent = {'value': cvalue, 'size': size, 'bold': bold, 'italic': italic,
                               'underline': underline,
                               'fontname': 'Arial', 'valign': VerticalAlignment07Dict.get(valign),
                               'halign': HorizontalAlignment07Dict.get(halign),
                               'tborder': Border07Dict.get(tborder), 'bborder': Border07Dict.get(bborder),
                               'lborder': Border07Dict.get(lborder), 'rborder': Border07Dict.get(rborder)}
                # store a cell's data and format
                rowcontent.append(cellcontent)
                jj += 1
            sheetcontent.append(rowcontent)
            # print()
            ii += 1
        # print('--------------------------------------------------')
        # print(type(cell).__name__)
        # print(cell.coordinate)
        merge = []  # store sheet's merged cells
        for m in sheet.merged_cells.ranges:
            # print(m.__dict__)
            merge.append(((m.min_row - 1, m.min_col - 1), (m.max_row - 1, m.max_col - 1)))
        # print('--------------------------------------------------')

        content.append([sheetcontent, sheetname, merge])
    test = cell.border.top
    # showDetail(test)
    # print('\n',content)
    return content


# function used for seeking methods or attributes
def showDetail(t):
    print('------This is test part------')
    print(str(t), ':', t)
    print('-----------------------------')
    print('dir():', dir(t))
    print('-----------------------------')
    print('type():', type(t))
    print('-----------------------------')
    try:
        print('dump():', t.dump)
    except Exception as e:
        print('Exception:', e)


if __name__ == '__main__':
    file_2003 = './2003.xls'
    file_2007 = './2007.xlsx'

    # read03xls(file_2003)
    # read07xlsx(file_2007)
    # readCSV('./cc.csv')
