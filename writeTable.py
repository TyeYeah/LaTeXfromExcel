import xlwt
import openpyxl
from readTable import *

# Below are dictionaries for data transform

ReHorizontalAlignment03Dict = {'Default': 0, 'Left': 1, 'Center': 2, 'Right': 3, 'Justified': 5, 'Filled': 4,
                               'Distributed': 7}
ReHorizontalAlignment07Dict = {'Default': None, 'Default': 'general', 'Left': 'left', 'Center': 'center',
                               'Right': 'right', 'Justified': 'justify', 'Filled': 'fill', 'Distributed': 'distributed'}
ReVerticalAlignment03Dict = {'Top': 0, 'Middle': 1, 'Bottom': 2, 'Distributed': 4}
ReVerticalAlignment07Dict = {'Top': 'top', 'Middle': 'center', 'Bottom': 'bottom', 'Distributed': 'distributed'}

ReBorder03Dict = {'No': 0, 'Solid': 1, 'Dotted': 4, 'Dashed': 3, 'DashDot': 9, 'DashDotDot': 11, 'DoubleThin': 6,
                  None: 1}

ReBorder07Dict = {'No': None, 'Solid': 'thin', 'Dotted': 'dotted', 'Dashed': 'dashed', 'DashDot': 'dashDot',
                  'DashDotDot': 'dashDotDot', 'DoubleThin': 'double'}
ReUnderline07Dit = {0: None, 1: 'single', None: None, 'single': 'single'}


# check the border value and return right output format
def borderCheck(border):
    if border == 'No':
        return ''
        pass
    else:
        return '|'


# check the alignment and return right output format
def alignCheck(align):
    if align == 'Left':
        return 'l'
    elif align == 'Right':
        return 'r'
    else:
        return 'c'


# pack cell data according to bold value
def boldPack(value, bold):
    if (bold == 'True') or (bold == 1):
        # print('\\textbf{' + value + '}')
        return '\\textbf{' + value + '}'
    else:
        # print(bold)
        return value


# pack cell data according to italic value
def italicPack(value, italic):
    if (italic == 'True') or (italic == 1):
        # print('\\textit{' + value + '}')
        return '\\textit{' + value + '}'
    else:
        # print(italic)
        return value


# pack cell data according to underline value
def underlinePack(value, underline):
    if (underline == 'single') or (underline == 1):
        # print('\\underline{' + value + '}')
        return '\\underline{' + value + '}'
    else:
        # print(underline)
        return value


# common usage for packing normal cell:
# underlinePack(boldPack(italicPack(col.get('value'),col.get('italic')),col.get('bold')),col.get('underline'))

# write LaTeX source to 'path', from 'content', which comes from my own data format
def writeLaTeX(path, content):
    source = ''
    with open(path, 'w') as f:
        # write macro
        f.write('\documentclass{ctexart}\n')
        f.write('\\usepackage{multirow}\n')
        f.write('\\begin{document}\n')
        f.write('\\begin{table}[]\n')  # write table head
    source+='\documentclass{ctexart}\n'+'\\usepackage{multirow}\n'+'\\begin{document}\n'+'\\begin{table}[]\n'
    for sheets in content:
        sheetname = sheets[1]
        merge = sheets[2]
        print('----------sheet:', sheetname, '----------')
        ii = jj = 0
        colnum = sheets[0][0].__len__()
        with open(path, 'a') as f:
            f.write('\\begin{tabular}{' + 'c' * colnum + '}\n')  # write tabular head
            source+='\\begin{tabular}{' + 'c' * colnum + '}\n'
        for row in sheets[0]:
            line = ''
            topline = ''
            bottomline = ''
            topborder = []
            bottomborder = []
            for col in row:
                flag = 0
                for mergeunit in merge:  # check cell type

                    if ii == mergeunit[0][0] and jj == mergeunit[0][1] and mergeunit[0][0] == mergeunit[1][0]:
                        flag = 1  # first cell of a merged cell with multiple column but one row
                        cellstr = ' \multicolumn{' + str(mergeunit[1][1] - mergeunit[0][1] + 1) + '}{' + borderCheck(
                            col.get('lborder')) + alignCheck(col.get('halign')) + borderCheck(
                            col.get('rborder')) + '}{' + underlinePack(
                            boldPack(italicPack(col.get('value'), col.get('italic')), col.get('bold')),
                            col.get('underline')) + '} '
                        line += cellstr

                        if col.get('tborder') != 'No':
                            topborder.append(jj)
                        if col.get('bborder') != 'No':
                            bottomborder.append(jj)

                        break
                        pass
                    elif ii == mergeunit[0][0] and jj == mergeunit[0][1] and mergeunit[0][1] == mergeunit[1][1]:
                        flag = 2  # first cell of a merged cell with multiple row but one column
                        if col.get('tborder') != 'No':
                            topborder.append(jj)
                        cellstr = ' \multirow{' + str(mergeunit[1][0] - mergeunit[0][0]) + '}{*}{' + underlinePack(
                            boldPack(italicPack(col.get('value'), col.get('italic')), col.get('bold')),
                            col.get('underline')) + '} '
                        if (jj + 1) == row.__len__():
                            cellstr += ' '
                        else:
                            cellstr += ' &'
                        line += cellstr
                        break
                        pass
                    elif ii == mergeunit[0][0] and jj == mergeunit[0][1] and mergeunit[0][1] < mergeunit[1][1] and \
                            mergeunit[0][0] < mergeunit[1][0]:
                        flag = 3# first cell of a merged cell with multiple rows and columns
                        if col.get('tborder') != 'No':
                            topborder.append(jj)
                        cellstr = ' \multicolumn{' + str(mergeunit[1][1] - mergeunit[0][1] + 1) + '}{' + borderCheck(
                            col.get('lborder')) + alignCheck(col.get('halign')) + borderCheck(
                            col.get('rborder')) + '}{\multirow{' + str(
                            mergeunit[1][0] - mergeunit[0][0] + 1) + '}{*}{' + underlinePack(
                            boldPack(italicPack(col.get('value'), col.get('italic')), col.get('bold')),
                            col.get('underline')) + '}} '
                        line += cellstr
                        break
                        pass
                    elif mergeunit[0][0] == mergeunit[1][0] and ii == mergeunit[0][0] and (
                            jj in range(mergeunit[0][1] + 1, mergeunit[1][1] + 1)):
                        flag = 4# not the first cell of a merged cell with multiple column but one row
                        if col.get('tborder') != 'No':
                            topborder.append(jj)
                        if col.get('bborder') != 'No':
                            bottomborder.append(jj)
                        if jj == mergeunit[1][1]:
                            cellstr = ' &'
                        else:
                            cellstr = ' '
                        line += cellstr
                        break
                        pass
                    elif mergeunit[0][1] == mergeunit[1][1] and jj == mergeunit[0][1] and (
                            ii in range(mergeunit[0][0] + 1, mergeunit[1][0] + 1)):
                        flag = 5# not the first cell of a merged cell with multiple row but one column
                        if ii == mergeunit[1][0] and sheets[0][mergeunit[0][0]][mergeunit[0][1]].get('bborder') != 'No':
                            bottomborder.append(jj)
                            pass
                        if (jj + 1) == row.__len__():
                            cellstr = ' '
                        else:
                            cellstr = ' &'
                        line += cellstr
                        break
                        pass
                    elif (ii in range(mergeunit[0][0], mergeunit[1][0] + 1)) and (
                            jj in range(mergeunit[0][1], mergeunit[1][1] + 1)) and jj == mergeunit[0][1]:
                        flag = 6#first column of a merged cell with multiple rows and columns
                        if ii == mergeunit[1][0] and sheets[0][mergeunit[0][0]][mergeunit[0][1]].get('bborder') != 'No':
                            bottomborder.append(jj)
                        cellstr = ' \multicolumn{' + str(mergeunit[1][1] - mergeunit[0][1] + 1) + '}{' + borderCheck(
                            sheets[0][mergeunit[0][0]][
                                mergeunit[0][1]].get('lborder')) + alignCheck(col.get('halign')) + borderCheck(
                            sheets[0][mergeunit[0][0]][mergeunit[0][1]].get(
                                'rborder')) + '}{}   '
                        line += cellstr
                        break
                        pass
                    elif (ii in range(mergeunit[0][0], mergeunit[1][0] + 1)) and (
                            jj in range(mergeunit[0][1], mergeunit[1][1] + 1)) and jj != mergeunit[0][1]:
                        flag = 7#not first column of a merged cell with multiple rows and columns
                        if ii == mergeunit[0][0] and sheets[0][mergeunit[0][0]][mergeunit[0][1]].get('tborder') != 'No':
                            topborder.append(jj)
                            pass
                        if ii == mergeunit[1][0] and sheets[0][mergeunit[0][0]][mergeunit[0][1]].get('bborder') != 'No':
                            bottomborder.append(jj)
                        if jj == mergeunit[1][1]:
                            cellstr = ' &'
                        else:
                            cellstr = ' '
                        line += cellstr
                        break
                        pass

                if flag == 0:#cells which are not belonging to merged cells
                    cellstr = '\multicolumn{1}{' + borderCheck(
                        col.get('lborder')) + alignCheck(col.get('halign')) + borderCheck(
                        col.get('rborder')) + '}{' + underlinePack(
                        boldPack(italicPack(col.get('value'), col.get('italic')), col.get('bold')),
                        col.get('underline')) + '}'
                    if (jj + 1) == row.__len__():
                        cellstr += ''
                    else:
                        cellstr += '&'
                    line += cellstr
                    if col.get('tborder') != 'No':
                        topborder.append(jj)
                    if col.get('bborder') != 'No':
                        bottomborder.append(jj)
                    pass
                else:
                    pass
                jj += 1
                pass
            line += '\\\\'

            for col in topborder:
                topline += '\cline{' + str(col + 1) + '-' + str(col + 1) + '}'
            for col in bottomborder:
                bottomline += '\cline{' + str(col + 1) + '-' + str(col + 1) + '}'

            with open(path, 'a') as f:
                f.write(topline + '\n')
                f.write(line + '\n')
                f.write(bottomline + '\n')
            source+=topline + '\n'+line + '\n'+bottomline + '\n'
            jj = 0
            ii += 1
        with open(path, 'a') as f:
            f.write('\end{tabular}\n')
        source+='\end{tabular}\n'
    with open(path, 'a') as f:
        f.write('\end{table}\n')
        f.write('\end{document}\n')
    source+='\end{table}\n'+'\end{document}\n'
    return source
    pass


# write HTML to 'path' from 'content'
def writeHTML(path, content):
    pass


# write xls file to 'path' from 'content'
def write03xls(path, content=[]):
    workbook = xlwt.Workbook()
    for sheets in content:
        sheetname = sheets[1]
        merge = sheets[2]
        print(sheetname)
        sheet = workbook.add_sheet(sheetname)
        ii = jj = 0
        for row in sheets[0]:
            for col in row:
                style = xlwt.XFStyle()
                style.alignment.vert = ReVerticalAlignment03Dict.get(col.get('valign'))
                style.alignment.horz = ReHorizontalAlignment03Dict.get(col.get('halign'))
                style.borders.top = ReBorder03Dict.get(col.get('tborder'))
                style.borders.bottom = ReBorder03Dict.get(col.get('bborder'))
                style.borders.left = ReBorder03Dict.get(col.get('lborder'))
                style.borders.right = ReBorder03Dict.get(col.get('rborder'))
                style.font.bold = int(bool(col.get('bold')))
                style.font.italic = int(bool(col.get('italic')))
                style.font.name = col.get('fontname')
                style.font.underline = int(bool(col.get('underline')))
                style.font.height = int(20 * col.get('size'))
                # print(style.font.height)

                flag = 0
                for mergeunit in merge:  # check cell type
                    if ii == mergeunit[0][0] and jj == mergeunit[0][1]:
                        flag = 1#belongs to merged cell with merged cell format
                        break
                        pass
                    elif (ii in range(mergeunit[0][0], mergeunit[1][0] + 1)) and (
                            jj in range(mergeunit[0][1], mergeunit[1][1] + 1)):
                        flag = 2#belongs to merged cell but have nothing to do with final cell format
                        break
                        pass

                if flag == 0:
                    sheet.write(ii, jj, col.get('value'), style)
                    pass
                elif flag == 1:
                    sheet.write_merge(mergeunit[0][0], mergeunit[1][0], mergeunit[0][1], mergeunit[1][1],
                                      col.get('value'), style)
                    pass
                elif flag == 2:
                    pass

                jj += 1
            jj = 0
            ii += 1

    workbook.save(path)
    pass


# write xlsx file to 'path' from 'content'
def write07xlsx(path, content=[]):
    workbook = openpyxl.Workbook()
    for sheets in content:
        sheet = workbook.active
        sheet.title = sheets[1]
        merge = sheets[2]
        print(sheet.title)

        ii = jj = 1
        for row in sheets[0]:
            for col in row:
                style = sheet.cell(ii, jj)
                style.value = col.get('value')
                style.alignment = openpyxl.styles.Alignment(vertical=ReVerticalAlignment07Dict.get(col.get('valign')),
                                                            horizontal=ReHorizontalAlignment07Dict.get(
                                                                col.get('halign')))
                top = ReBorder07Dict.get(col.get('tborder'))
                topside = openpyxl.styles.borders.Side(top)
                bottom = ReBorder07Dict.get(col.get('bborder'))
                bottomside = openpyxl.styles.borders.Side(bottom)
                left = ReBorder07Dict.get(col.get('lborder'))
                leftside = openpyxl.styles.borders.Side(left)
                right = ReBorder07Dict.get(col.get('rborder'))
                rightside = openpyxl.styles.borders.Side(right)
                style.border = openpyxl.styles.Border(top=topside, bottom=bottomside, left=leftside, right=rightside)
                style.font = openpyxl.styles.Font(bold=bool(int(col.get('bold'))), italic=bool(int(col.get('italic'))),
                                                  name=col.get('fontname'),
                                                  underline=ReUnderline07Dit.get(col.get('underline')),
                                                  size=col.get('size'))

                jj += 1
            jj = 1
            ii += 1

        for mergeunit in merge:#merge cells at last
            mergeexp = (chr(mergeunit[0][1] + ord('A')) + str(mergeunit[0][0] + 1) + ':' + chr(
                mergeunit[1][1] + ord('A')) + str(mergeunit[1][0] + 1))
            print(mergeexp)
            sheet.merge_cells(mergeexp)
    workbook.save(path)
    pass


# write csv file to 'path' from 'content'
def writeCSV(path, content):
    with open(path, 'w') as f:
        for sheet in content:
            ii = jj = 0
            for row in sheet[0]:
                for col in row:
                    f.write(col.get('value'))
                    if jj + 1 == row.__len__():
                        f.write('\n')
                    else:
                        f.write(',')
                    jj += 1
                jj = 0
                ii += 1
    pass


if __name__ == '__main__':
    file_2003 = './2003.xls'
    file_2007 = './2007.xlsx'
    write_file_2003 = './write_2003.xls'
    write_file_2007 = './write_2007.xlsx'

    # write03xls(write_file_2003, read07xlsx(file_2007))
    # styleee = xlwt.XFStyle()
    # print(dir(styleee))
    # patternnn = xlwt.Pattern()
    # print(dir(patternnn))
    # print(dir(styleee.font))  # bold, italic, height, name, underline
    # print(dir(styleee.borders))  # top, bottom, left, right
    # print(dir(styleee.alignment))  # vert, horz

    # write07xlsx(write_file_2007, read03xls(file_2003))
    # cell = openpyxl.Workbook().active.cell(1, 1, 123456)
    # print(cell.font)
    # print(dir(cell))

    # writeLaTeX('./test.tex', readCSV('./cc.csv'))
    # writeCSV('./ccc.csv',readCSV('./cc.csv'))
